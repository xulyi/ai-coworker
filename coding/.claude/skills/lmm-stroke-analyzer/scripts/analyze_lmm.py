#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
脑卒中康复纵向数据 LMM 分析与目标达成检查
用法: python analyze_lmm.py <数据文件.xlsx|.csv> [--output-dir <输出目录>]

输出:
  1. Markdown 表格到 stdout
  2. Excel 文件到输出目录
"""

import argparse
import os
import sys
import warnings

import numpy as np
import pandas as pd
from scipy import stats
import statsmodels.formula.api as smf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

warnings.filterwarnings('ignore')

# ============================================================================
# 配置
# ============================================================================
GROUP_COL = '分组'
ID_COL = '患者ID'
TIME_COL = '时间点'
TIME_ORDER = ['T0', 'T1', 'T2', 'T3']
GROUP_ORDER = [1, 2, 3, 4]

OUTCOMES = {
    'FMA_LE': 'higher_better',
    'ADL':    'higher_better',
    'BBS':    'higher_better',
    'TUGT':   'lower_better',
    'CSS':    'lower_better',
    'MAS':    'lower_better',
}

POSTHOC_TIMES = ['T2', 'T3']
REF_GROUP = 1
REF_TIME = 'T0'
BASELINE_P_THRESHOLD = 0.05

# 目标对比定义（参数 -> 时点 -> [(显示名, g_a, g_b, 目标类型), ...]）
TARGET_CONTRASTS = {
    'FMA_LE': {
        'T2': [('G1>G2', 1, 2, 'sig_moderate'), ('G2>G3', 2, 3, 'sig_any'), ('G3=G4', 3, 4, 'nonsig_loose')],
        'T3': [('G1>G2', 1, 2, 'sig_moderate'), ('G2>G3', 2, 3, 'sig_any'), ('G3=G4', 3, 4, 'nonsig_loose')],
    },
    'ADL': {
        'T2': [('G1>G2', 1, 2, 'sig_moderate'), ('G2>G3', 2, 3, 'sig_any'), ('G3=G4', 3, 4, 'nonsig_loose')],
        'T3': [('G1>G2', 1, 2, 'nonsig_loose'), ('G2>G3', 2, 3, 'sig_any'), ('G3=G4', 3, 4, 'nonsig_loose')],
    },
    'BBS': {
        'T2': [('G1>G2', 1, 2, 'sig_moderate'), ('G2>G3', 2, 3, 'sig_any'), ('G3=G4', 3, 4, 'nonsig_loose')],
        'T3': [('G1>G2', 1, 2, 'nonsig_loose'), ('G2>G3', 2, 3, 'sig_any'), ('G3=G4', 3, 4, 'nonsig_loose')],
    },
    'MAS': {
        'T2': [('G1>G2', 1, 2, 'sig_moderate'), ('G2>G3', 2, 3, 'sig_any'), ('G3=G4', 3, 4, 'nonsig_loose')],
        'T3': [('G1>G2', 1, 2, 'nonsig_loose'), ('G2>G3', 2, 3, 'sig_any'), ('G3=G4', 3, 4, 'nonsig_loose')],
    },
    'CSS': {
        'T2': [('G1>G2', 1, 2, 'sig_moderate'), ('G2>G3', 2, 3, 'sig_any'), ('G3=G4', 3, 4, 'nonsig_loose')],
        'T3': [('G1>G2', 1, 2, 'nonsig_loose'), ('G2>G3', 2, 3, 'sig_any'), ('G3=G4', 3, 4, 'nonsig_loose')],
    },
}

# ============================================================================
# 工具函数
# ============================================================================

def load_data(path):
    if path.lower().endswith('.csv'):
        df = pd.read_csv(path)
    elif path.lower().endswith(('.xlsx', '.xls')):
        df = pd.read_excel(path)
    else:
        raise ValueError('仅支持 .csv 或 .xlsx 格式')
    return clean_df(df)


def clean_df(df):
    df = df.copy()
    df.columns = df.columns.str.strip().str.replace('\ufeff', '', regex=False)
    df[ID_COL] = df[ID_COL].astype(str).str.strip()
    df[TIME_COL] = df[TIME_COL].astype(str).str.strip().str.upper()
    df[GROUP_COL] = pd.to_numeric(df[GROUP_COL], errors='coerce').astype('Int64')
    df = df[df[GROUP_COL].isin(GROUP_ORDER)].copy()
    df[GROUP_COL] = pd.Categorical(df[GROUP_COL].astype(int), categories=GROUP_ORDER, ordered=True)
    df[TIME_COL] = pd.Categorical(df[TIME_COL], categories=TIME_ORDER, ordered=True)
    return df


def safe_desc(x):
    x = pd.to_numeric(x, errors='coerce').dropna()
    if len(x) == 0:
        return np.nan, np.nan, 0
    if len(x) == 1:
        return x.mean(), np.nan, len(x)
    return x.mean(), x.std(ddof=1), len(x)


def baseline_anova(df, metric):
    d = df[df[TIME_COL] == REF_TIME][[GROUP_COL, metric]].dropna().copy()
    groups = [pd.to_numeric(d[d[GROUP_COL] == g][metric], errors='coerce').dropna().values for g in GROUP_ORDER]
    if any(len(x) == 0 for x in groups):
        return {'metric': metric, 'anova_p': 0.0, 'merge_p': 0.0, 'aligned': False}
    try:
        _, p_anova = stats.f_oneway(*groups)
    except Exception:
        p_anova = 0.0
    try:
        g12 = np.concatenate([groups[0], groups[1]])
        g34 = np.concatenate([groups[2], groups[3]])
        _, p_merge = stats.ttest_ind(g12, g34, equal_var=False)
    except Exception:
        p_merge = 0.0
    return {
        'metric': metric, 'anova_p': p_anova, 'merge_p': p_merge,
        'aligned': p_anova >= BASELINE_P_THRESHOLD and p_merge >= BASELINE_P_THRESHOLD
    }


def find_param(index, group_val=None, time_val=None):
    g_part = None if group_val is None else f'[T.{group_val}]'
    t_part = None if time_val is None else f'[T.{time_val}]'
    matches = []
    for name in index:
        ok = True
        if g_part is not None and g_part not in name:
            ok = False
        if t_part is not None and t_part not in name:
            ok = False
        if 'C(' + GROUP_COL not in name and g_part is not None:
            ok = False
        if 'C(' + TIME_COL not in name and t_part is not None:
            ok = False
        if ok:
            matches.append(name)
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        return min(matches, key=len)
    return None


def linear_contrast(result, coef_names, weights):
    params = result.params
    cov = result.cov_params()
    est = 0.0
    valid = [(w, nm) for w, nm in zip(weights, coef_names) if nm is not None]
    for w, nm in valid:
        est += w * params[nm]
    var = 0.0
    for i, (wi, ni) in enumerate(valid):
        for j, (wj, nj) in enumerate(valid):
            var += wi * wj * cov.loc[ni, nj]
    var = float(var)
    se = np.sqrt(var) if var >= 0 else np.nan
    ci_low = est - 1.96 * se if pd.notna(se) else np.nan
    ci_high = est + 1.96 * se if pd.notna(se) else np.nan
    try:
        idx = result.params.index
        contrast = pd.Series(0.0, index=idx)
        for w, nm in valid:
            contrast[nm] = w
        R = contrast.values.reshape(1, -1)
        wald = result.wald_test(R, use_f=True, scalar=True)
        p = float(wald.pvalue)
    except Exception:
        z = est / se if se and se > 0 else np.nan
        p = 2 * (1 - stats.norm.cdf(abs(z))) if pd.notna(z) else np.nan
    return est, se, p, ci_low, ci_high


def pairwise_at_time(result, g_a, g_b, time_val):
    idx = result.params.index
    def main_effect(g):
        return None if g == REF_GROUP else find_param(idx, group_val=g, time_val=None)
    def interaction(g, t):
        return None if g == REF_GROUP or t == REF_TIME else find_param(idx, group_val=g, time_val=t)
    names = [main_effect(g_a), interaction(g_a, time_val), main_effect(g_b), interaction(g_b, time_val)]
    weights = [1, 1, -1, -1]
    if g_a == REF_GROUP:
        weights[0] = 0
        weights[1] = 0
    if g_b == REF_GROUP:
        weights[2] = 0
        weights[3] = 0
    return linear_contrast(result, names, weights)


def fit_lmm(formula, data, group_series):
    for method in ['lbfgs', 'powell', 'cg', 'nm']:
        try:
            model = smf.mixedlm(formula, data=data, groups=group_series, re_formula='~1')
            result = model.fit(reml=True, method=method, disp=False)
            return result, method, None
        except Exception as e:
            last_err = f'{type(e).__name__}: {e}'
    return None, None, last_err


def run_one_metric_lmm(df, metric):
    cols = [ID_COL, GROUP_COL, TIME_COL, metric]
    d = df[cols].dropna().copy()
    if d.empty:
        return None, None
    d[GROUP_COL] = pd.Categorical(d[GROUP_COL], categories=GROUP_ORDER, ordered=True)
    d[TIME_COL] = pd.Categorical(d[TIME_COL], categories=TIME_ORDER, ordered=True)
    formula = f"{metric} ~ C({GROUP_COL}, Treatment({REF_GROUP})) * C({TIME_COL}, Treatment('{REF_TIME}'))"
    result, method, err = fit_lmm(formula, d, d[ID_COL])
    if result is None:
        return None, None
    pair_rows = []
    pairs = [(1, 2), (1, 3), (1, 4), (2, 3), (2, 4), (3, 4)]
    for t in POSTHOC_TIMES:
        for g_a, g_b in pairs:
            est, se, p, ci_l, ci_h = pairwise_at_time(result, g_a, g_b, t)
            true_direction = OUTCOMES.get(metric, 'higher_better')
            if est > 0:
                raw_dir = f'G{g_a} > G{g_b}'
                clin_dir = 'G{}更优'.format(g_a if true_direction == 'higher_better' else g_b)
            elif est < 0:
                raw_dir = f'G{g_a} < G{g_b}'
                clin_dir = 'G{}更优'.format(g_b if true_direction == 'higher_better' else g_a)
            else:
                raw_dir = f'G{g_a} = G{g_b}'
                clin_dir = '无差异'
            pair_rows.append({
                'metric': metric, 'time': t, 'contrast': f'G{g_a}_vs_G{g_b}',
                'estimate': est, 'se': se, 'p': p,
                'ci95_low': ci_l, 'ci95_high': ci_h,
                'raw_direction': raw_dir, 'clinical_direction': clin_dir,
            })
    return result, pd.DataFrame(pair_rows)


def evaluate_cell(p, est, target_type, true_direction):
    if target_type == 'sig_moderate':
        dir_ok = (true_direction == 'higher_better' and est > 0) or (true_direction == 'lower_better' and est < 0)
        if not dir_ok:
            return '✗', 'FAIL', '方向错误'
        if 0.01 <= p < 0.05:
            return '✓', 'ideal', ''
        elif p < 0.01:
            return '✓', 'OK', '偏强'
        else:
            return '✗', 'FAIL', '不显著'
    elif target_type == 'sig_any':
        dir_ok = (true_direction == 'higher_better' and est > 0) or (true_direction == 'lower_better' and est < 0)
        if not dir_ok:
            return '✗', 'FAIL', '方向错误'
        if 0.01 <= p < 0.05:
            return '✓', 'ideal', ''
        elif p < 0.01:
            return '✓', 'OK', '偏强'
        else:
            return '✗', 'FAIL', '不显著'
    else:  # nonsig_loose
        if p >= 0.05:
            return '✓', 'ideal', ''
        elif p >= 0.03:
            return '✓', 'OK', '边缘'
        else:
            return '✗', 'FAIL', '过显著'


def sig_symbol(p):
    if p < 0.001:
        return '***'
    elif p < 0.01:
        return '**'
    elif p < 0.05:
        return '*'
    else:
        return 'ns'


# ============================================================================
# 分析主流程
# ============================================================================

def analyze(df, output_dir, export_excel=True):
    # 1. T0 基线
    baseline_rows = []
    for metric in OUTCOMES.keys():
        baseline_rows.append(baseline_anova(df, metric))

    # 2. 描述统计
    desc_data = {}
    for metric in OUTCOMES.keys():
        desc_data[metric] = {}
        for time in TIME_ORDER:
            desc_data[metric][time] = {}
            for g in GROUP_ORDER:
                dsub = df[(df[GROUP_COL] == g) & (df[TIME_COL] == time)][metric]
                m, s, n = safe_desc(dsub)
                desc_data[metric][time][g] = (m, s, n)

    # 3. LMM
    lmm_results = {}
    for metric in OUTCOMES.keys():
        result, pair_df = run_one_metric_lmm(df, metric)
        lmm_results[metric] = pair_df

    # 4. 目标达成总表
    summary_rows = []
    for metric in ['FMA_LE', 'ADL', 'BBS', 'MAS', 'CSS']:
        for t in ['T2', 'T3']:
            pair_df = lmm_results.get(metric)
            if pair_df is None:
                continue
            true_direction = OUTCOMES[metric]
            row = {'参数': metric, '时点': t}
            statuses = []
            concerns = []

            for col_name, g_a, g_b, target_type in TARGET_CONTRASTS[metric][t]:
                sub = pair_df[(pair_df['time'] == t) & (pair_df['contrast'] == f'G{g_a}_vs_G{g_b}')]
                if sub.empty:
                    row[col_name] = '无数据'
                    statuses.append('fail')
                    concerns.append(f'{col_name}: 无数据')
                    continue
                p = sub['p'].values[0]
                est = sub['estimate'].values[0]
                icon, label, detail = evaluate_cell(p, est, target_type, true_direction)
                sig = sig_symbol(p)
                cell_text = f'P={p:.4f} {sig} {icon}\n{label}'
                row[col_name] = cell_text
                if icon == '✗':
                    statuses.append('fail')
                    concerns.append(f'{col_name}: {detail}')
                elif label == 'OK' and detail:
                    statuses.append('ok')
                    concerns.append(f'{col_name}: {detail}')
                else:
                    statuses.append('ideal')

            if 'fail' in statuses or 'ok' in statuses:
                # 按 detail 分组，同 detail 的对比名用顿号连接，格式：⚠️ G1>G2、G2>G3 未显著
                from collections import defaultdict
                detail_groups = defaultdict(list)
                for c in concerns:
                    name, d = c.split(': ', 1)
                    d = d.replace('不显著', '未显著')
                    detail_groups[d].append(name)
                parts = []
                for d, names in detail_groups.items():
                    parts.append('、'.join(names) + ' ' + d)
                row['状态'] = '⚠️ ' + '、'.join(parts)
            else:
                row['状态'] = '✅ 全部达标'

            summary_rows.append(row)

    summary_df = pd.DataFrame(summary_rows)

    # ========================================================================
    # 输出 Markdown
    # ========================================================================
    print("=" * 80)
    print("LMM 目标达成总表")
    print("=" * 80)
    print()
    print("| 参数 | 时点 | G1>G2 | G2>G3 | G3=G4 | 状态 |")
    print("|------|------|-------|-------|-------|------|")
    for _, row in summary_df.iterrows():
        cells = [row['参数'], row['时点']]
        for col in ['G1>G2', 'G2>G3', 'G3=G4']:
            val = str(row[col]).replace('\n', '<br>')
            cells.append(val)
        status = str(row['状态']).replace('\n', '<br>')
        cells.append(status)
        print("| " + " | ".join(cells) + " |")

    print()
    print("=" * 80)
    print("二、所有参数 × 所有时间点 × 各组 均值 ± 标准差")
    print("=" * 80)
    for metric in OUTCOMES.keys():
        print()
        print(f"**{metric}**")
        print()
        print("| 时点 | G1 | G2 | G3 | G4 |")
        print("|------|-----|-----|-----|-----|")
        for time in TIME_ORDER:
            cells = [time]
            for g in GROUP_ORDER:
                m, s, n = desc_data[metric][time][g]
                if pd.isna(m):
                    cells.append('-')
                elif pd.isna(s) or n <= 1:
                    cells.append(f"{m:.2f}±NA(n={n})")
                else:
                    cells.append(f"{m:.2f}±{s:.2f}")
            print("| " + " | ".join(cells) + " |")

    print()
    print("=" * 80)
    print("三、T0 基线 P值汇总")
    print("=" * 80)
    print()
    print("| 参数 | ANOVA P | Merge(G1+2 vs G3+4) P | 状态 |")
    print("|------|---------|----------------------|------|")
    for r in baseline_rows:
        status = "✅ 对齐" if r['aligned'] else "❌ 未对齐"
        print(f"| {r['metric']} | {r['anova_p']:.4f} | {r['merge_p']:.4f} | {status} |")

    # ========================================================================
    # 导出 Excel
    # ========================================================================
    if not export_excel:
        return

    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()

    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    subheader_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
    subheader_font = Font(color='FFFFFF', bold=True, size=10)
    green_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
    red_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    yellow_fill = PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def apply_border(ws, start_row, start_col, end_row, end_col):
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Sheet 1
    ws1 = wb.active
    ws1.title = "LMM目标达成总表"
    ws1.merge_cells('A1:F1')
    ws1['A1'] = 'LMM 目标达成总表'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30

    hdr = ['参数', '时点', 'G1>G2', 'G2>G3', 'G3=G4', '状态']
    for c, h in enumerate(hdr, 1):
        cell = ws1.cell(row=2, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for r_idx, (_, row) in enumerate(summary_df.iterrows(), 3):
        ws1.cell(row=r_idx, column=1, value=row['参数'])
        ws1.cell(row=r_idx, column=2, value=row['时点'])
        ws1.cell(row=r_idx, column=3, value=row['G1>G2'])
        ws1.cell(row=r_idx, column=4, value=row['G2>G3'])
        ws1.cell(row=r_idx, column=5, value=row['G3=G4'])
        ws1.cell(row=r_idx, column=6, value=row['状态'])
        apply_border(ws1, r_idx, 1, r_idx, 6)
        status = str(row['状态'])
        if '全部达标' in status:
            fill = green_fill
        elif '未达标' in status:
            fill = red_fill
        else:
            fill = yellow_fill
        for c in range(1, 7):
            ws1.cell(row=r_idx, column=c).fill = fill

    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 8
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 22
    ws1.column_dimensions['F'].width = 35

    # 描述统计 sheets
    for metric in OUTCOMES.keys():
        ws = wb.create_sheet(title=metric)
        ws.merge_cells('A1:E1')
        ws['A1'] = f'{metric} - 均值 ± 标准差'
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        h = ['时点', 'G1', 'G2', 'G3', 'G4']
        for c, v in enumerate(h, 1):
            cell = ws.cell(row=2, column=c, value=v)
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for r_idx, time in enumerate(TIME_ORDER, 3):
            ws.cell(row=r_idx, column=1, value=time)
            for g in GROUP_ORDER:
                m, s, n = desc_data[metric][time][g]
                if pd.isna(m):
                    val = '-'
                elif pd.isna(s) or n <= 1:
                    val = f"{m:.2f}±NA"
                else:
                    val = f"{m:.2f}±{s:.2f}"
                ws.cell(row=r_idx, column=g + 1, value=val)
            apply_border(ws, r_idx, 1, r_idx, 5)

        for c in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[c].width = 16

    # T0 基线 sheet
    ws_t0 = wb.create_sheet(title='T0基线')
    ws_t0.merge_cells('A1:D1')
    ws_t0['A1'] = 'T0 基线组间比较'
    ws_t0['A1'].font = Font(bold=True, size=12)
    ws_t0['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_t0.row_dimensions[1].height = 25

    h = ['参数', 'ANOVA P', 'Merge P (G1+2 vs G3+4)', '状态']
    for c, v in enumerate(h, 1):
        cell = ws_t0.cell(row=2, column=c, value=v)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r_idx, r in enumerate(baseline_rows, 3):
        ws_t0.cell(row=r_idx, column=1, value=r['metric'])
        ws_t0.cell(row=r_idx, column=2, value=f"{r['anova_p']:.4f}")
        ws_t0.cell(row=r_idx, column=3, value=f"{r['merge_p']:.4f}")
        status = '对齐' if r['aligned'] else '未对齐'
        ws_t0.cell(row=r_idx, column=4, value=status)
        apply_border(ws_t0, r_idx, 1, r_idx, 4)
        fill = green_fill if r['aligned'] else red_fill
        for c in range(1, 5):
            ws_t0.cell(row=r_idx, column=c).fill = fill

    for c in ['A', 'B', 'C', 'D']:
        ws_t0.column_dimensions[c].width = 22

    excel_path = os.path.join(output_dir, 'LMM分析结果.xlsx')
    wb.save(excel_path)
    print()
    print(f"Excel 已导出: {excel_path}")


# ============================================================================
# 入口
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description='脑卒中康复纵向数据 LMM 分析与目标达成检查')
    parser.add_argument('input', help='输入数据文件 (.csv 或 .xlsx)')
    parser.add_argument('--output-dir', '-o', default='.', help='输出目录 (默认当前目录)')
    parser.add_argument('--no-excel', action='store_true', help='不导出 Excel，仅屏幕输出')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"错误: 找不到文件 {args.input}")
        sys.exit(1)

    df = load_data(args.input)
    print(f"数据加载完成: {len(df)} 条记录, {df[ID_COL].nunique()} 名患者")
    analyze(df, args.output_dir, export_excel=not args.no_excel)


if __name__ == '__main__':
    main()
